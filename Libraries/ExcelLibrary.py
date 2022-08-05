"""Module file containing class and functions to read and write data in Excel"""

from openpyxl import load_workbook
from robot.api.deco import keyword
from os.path import exists


class ExcelLibrary:
    ROBOT_LIBRARY_SCOPE = 'TEST'
    ROBOT_AUTO_KEYWORDS = False
    ROBOT_LISTENER_API_VERSION = 3

    def __init__(self):
        self.ROBOT_LIBRARY_LISTENER = self
        self.test_name = ""
        self.excel_file_path = ""

    # noinspection PyUnusedLocal
    def start_test(self, data, result):
        """Listener implementation to set the name of the test for reading data"""
        self.test_name = data.name

    @keyword
    def set_excel_file_path(self, excel_file_path):
        """Function to set file path of the Excel workbook

        param excel_file_path: Path of the Excel workbook
        """
        if exists(excel_file_path):
            self.excel_file_path = excel_file_path
            print("Specified path is set as data source path - " + excel_file_path)
        else:
            raise Exception("File is not available in the specified path - " + excel_file_path)

    @classmethod
    def get_cell_value(cls, cell):
        """Function to get value from the Excel cell

        param cell: Cell object from Sheet object of openpyxl library
        :return: text value in the cell
        """
        cell_value = cell.value
        if cell_value is None:
            cell_value = ""
        return cell_value

    @keyword
    def get_data_from_cell(self, sheet_name, cell_name):
        """Function to get data from Excel cell

        param sheet_name: name of the sheet to get data
        param cell_name: cell identifier e.g. D2
        return: text value in the cell
        """
        workbook = load_workbook(self.excel_file_path)
        try:
            self.check_sheet_exists(workbook, sheet_name)
            cell_value = self.get_cell_value(workbook[sheet_name][cell_name])
            return cell_value
        finally:
            workbook.close()

    @keyword
    def put_data_in_cell(self, sheet_name, cell_name, cell_value):
        """Function to set data in Excel cell

        param sheet_name: name of the sheet to get data
        param cell_name: cell identifier e.g. D2
        param cell_value: value to be stored
        """
        workbook = load_workbook(self.excel_file_path)
        try:
            self.check_sheet_exists(workbook, sheet_name)
            workbook[sheet_name][cell_name].value = cell_value
            workbook.save(self.excel_file_path)
            print("Specified value - " + str(cell_value) + " is set in the cell - " + cell_name)
        finally:
            workbook.close()

    @classmethod
    def check_sheet_exists(cls, workbook, sheet_name):
        """Class Method to check desired sheet exists in workbook

        param workbook: workbook object of openpyxl library
        param sheet_name: name of the sheet to check
        """
        if sheet_name not in workbook.get_sheet_names():
            raise Exception("Sheet with name " + sheet_name + " does not exists")

    def get_column_index(self, sheet, column_name):
        """Function to get column index from the spreadsheet

        param sheet: name of the sheet to check the specified column
        param column_name: column name for which index has to be determined
        return: index of the desired column
        """
        column_max_count = sheet.max_column + 1
        for index in range(1, column_max_count):
            cell_value = self.get_cell_value(sheet.cell(1, index))
            if cell_value == column_name:
                return index
        raise Exception("Specified column " + column_name + " is not found")

    def get_row_index(self, sheet):
        """Function to get index of row with matching Test_Name from specified sheet

        param sheet: name of the sheet from which index has to be found
        return: index of the desired row
        """
        column_index_test_name = self.get_column_index(sheet, "Test_Name")
        for index in range(2, sheet.max_row + 1):
            if self.get_cell_value(sheet.cell(index, column_index_test_name)) == self.test_name:
                return index
        raise Exception("Row with test name as " + self.test_name + " is not found in sheet - " + sheet.title)

    @keyword
    def get_data(self, sheet_name, column_name):
        """Function to get data from Excel

        param sheet_name: name of the sheet from which data is to be retrieved
        param column_name: name of the column from which data is to be retrieved
        return: text value present in the cell
        """
        workbook = load_workbook(self.excel_file_path)
        try:
            self.check_sheet_exists(workbook, sheet_name)
            sheet = workbook[sheet_name]
            column_index_column = self.get_column_index(sheet, column_name)
            row_index = self.get_row_index(sheet)
            cell_value = self.get_cell_value(sheet.cell(row_index, column_index_column))
            return cell_value
        finally:
            workbook.close()

    @keyword
    def put_data(self, sheet_name, column_name, column_value):
        """Function to put data in Excel

        param sheet_name: name of the sheet in which data has to be stored
        param column_name: name of the column in which data has to be written
        param column_value: value which has to be written
        """
        workbook = load_workbook(self.excel_file_path)
        try:
            self.check_sheet_exists(workbook, sheet_name)
            sheet = workbook[sheet_name]
            column_index_column = self.get_column_index(sheet, column_name)
            row_index = self.get_row_index(sheet)
            sheet.cell(row_index, column_index_column).value = column_value
            workbook.save(self.excel_file_path)
            print("Specified value - " + str(column_value) + " is set in the column - " + column_name)
        finally:
            workbook.close()
